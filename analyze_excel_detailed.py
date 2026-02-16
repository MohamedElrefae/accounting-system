#!/usr/bin/env python3
"""
Detailed Excel File Analyzer
Properly analyzes complex Excel files with merged cells and multiple sheets
"""

import pandas as pd
import json
from pathlib import Path
import openpyxl

def analyze_excel_sheets(file_path):
    """Analyze all sheets in Excel file"""
    print(f"Analyzing Excel file: {file_path}")
    
    try:
        # Load Excel file with openpyxl to see sheet structure
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        print(f"\n📋 Available Sheets: {sheet_names}")
        
        for sheet_name in sheet_names:
            print(f"\n" + "="*60)
            print(f"Analyzing Sheet: '{sheet_name}'")
            print("="*60)
            
            # Read sheet with pandas
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                print(f"Sheet dimensions: {df.shape}")
                
                # Show first few rows to understand structure
                print(f"\n📊 First 15 rows:")
                for i in range(min(15, len(df))):
                    row_data = []
                    for j in range(min(10, len(df.columns))):
                        val = df.iloc[i, j]
                        if pd.isna(val):
                            row_data.append("NULL")
                        else:
                            row_data.append(str(val)[:20])  # Truncate long values
                    print(f"Row {i+1:2d}: {row_data}")
                
                # Try to identify header row
                print(f"\n🔍 Looking for headers...")
                for header_row in [0, 1, 2, 3, 4]:
                    if header_row < len(df):
                        print(f"Row {header_row+1} as headers: {list(df.iloc[header_row, :10])}")
                
                # Try reading with different header assumptions
                for header_row in [0, 1, 2, 3, 4]:
                    try:
                        df_with_headers = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                        print(f"\n📋 With header row {header_row}:")
                        print(f"Columns: {list(df_with_headers.columns)[:10]}")
                        
                        # Show sample data
                        if len(df_with_headers) > 0:
                            print(f"Sample data:")
                            print(df_with_headers.head(3).to_string())
                        
                        # Look for financial columns
                        cols = df_with_headers.columns.tolist()
                        financial_cols = []
                        for col in cols:
                            col_str = str(col).lower()
                            if any(keyword in col_str for keyword in ['debit', 'credit', 'مدين', 'دائن', 'amount', 'قيمة']):
                                financial_cols.append(col)
                        
                        if financial_cols:
                            print(f"\n💰 Financial columns found: {financial_cols}")
                            for col in financial_cols:
                                if col in df_with_headers.columns:
                                    col_data = df_with_headers[col].dropna()
                                    if len(col_data) > 0:
                                        print(f"  {col}: {len(col_data)} values, sum={col_data.sum():,.2f}")
                        
                        break  # Stop after first successful read
                    except Exception as e:
                        continue
                
            except Exception as e:
                print(f"Error reading sheet {sheet_name}: {e}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")

def analyze_transaction_data(file_path):
    """Try to find transaction data in Excel file"""
    print(f"\n🔍 Looking for transaction data patterns...")
    
    try:
        # Try different sheets and header combinations
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        for sheet_name in wb.sheetnames:
            print(f"\n--- Checking Sheet: {sheet_name} ---")
            
            # Read raw data
            ws = wb[sheet_name]
            data_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx > 100:  # Limit to first 100 rows for analysis
                    break
                
                # Filter out completely empty rows
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    data_rows.append((row_idx, row))
            
            print(f"Found {len(data_rows)} non-empty rows")
            
            # Look for patterns
            for i, (row_idx, row) in enumerate(data_rows[:20]):
                print(f"Row {row_idx}: {[str(cell)[:15] if cell is not None else 'NULL' for cell in row[:8]]}")
            
            # Try to identify transaction patterns
            print(f"\n🔍 Transaction Pattern Analysis:")
            for i, (row_idx, row) in enumerate(data_rows):
                # Look for numeric values (amounts)
                numeric_values = []
                for cell in row:
                    if cell is not None:
                        try:
                            val = float(cell)
                            if val != 0:
                                numeric_values.append(val)
                        except:
                            pass
                
                if numeric_values:
                    print(f"Row {row_idx}: Found amounts {numeric_values}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error in transaction analysis: {e}")

def main():
    """Main analysis function"""
    excel_files = [
        "c:/5/accounting-systemr5/يومية الحدائق من البداية كاملة .xlsx",
        "c:/5/accounting-systemr5/daily journal .xlsx",
        "c:/5/accounting-systemr5/يومية الحدائق من البداية حتى  12-2025.xlsx"
    ]
    
    for file_path in excel_files:
        if Path(file_path).exists():
            analyze_excel_sheets(file_path)
            analyze_transaction_data(file_path)
            break
        else:
            print(f"❌ File not found: {file_path}")

if __name__ == "__main__":
    main()
