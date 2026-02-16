#!/usr/bin/env python3
import openpyxl
import pandas as pd
import os

os.chdir(r'C:\5\accounting-systemr5')
excel_file = 'transactions.xlsx'

try:
    # First, let's see what sheets exist
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    print("Available sheets:", wb.sheetnames)
    
    # Try reading with pandas to see columns
    for sheet in wb.sheetnames:
        if 'transaction' in sheet.lower():
            print(f"\nReading sheet: '{sheet}'")
            df = pd.read_excel(excel_file, sheet_name=sheet, nrows=3)
            print(f"Columns ({len(df.columns)}):")
            for i, col in enumerate(df.columns, 1):
                print(f"  {i}: {col}")
            print(f"\nFirst row data:")
            for col in df.columns:
                print(f"  {col}: {df.iloc[0][col]}")
            break
            
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
