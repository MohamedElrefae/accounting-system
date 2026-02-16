#!/usr/bin/env python3
"""
Phase 0 Task 0.2: Excel Structure Validation

Verifies the Excel file structure, validates all expected columns are present,
checks data quality, and generates structure reports.
"""

import os
import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Tuple
from datetime import datetime
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

try:
    import pandas as pd
    from openpyxl import load_workbook
    from dotenv import load_dotenv
except ImportError as e:
    logger.error(f"Missing required package: {e}")
    logger.error("Install with: pip install pandas openpyxl python-dotenv")
    sys.exit(1)

load_dotenv()

EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH')

# Expected columns in Excel (from design document)
EXPECTED_COLUMNS = {
    'العام المالى': 'fiscal_year',
    'الشهر': 'month',
    'entry no': 'entry_no',
    'entry date': 'entry_date',
    'account code': 'account_code',
    'account name': 'account_name',
    'transaction classification code': 'transaction_classification_code',
    'classification code': 'classification_code',
    'classification name': 'classification_name',
    'project code': 'project_code',
    'project name': 'project_name',
    'work analysis code': 'work_analysis_code',
    'work analysis name': 'work_analysis_name',
    'sub_tree code': 'sub_tree_code',
    'sub_tree name': 'sub_tree_name',
    'مدين': 'debit',
    'دائن': 'credit',
    'ملاحظات': 'notes'
}


class ExcelStructureValidator:
    """Validates Excel file structure and data quality."""
    
    def __init__(self, file_path: str):
        """Initialize validator with Excel file path."""
        self.file_path = file_path
        self.workbook = None
        self.sheet_names = []
        self.data = None
        self.structure_info = {
            'timestamp': datetime.now().isoformat(),
            'file_path': file_path,
            'file_exists': False,
            'sheets': {},
            'validation_results': {},
            'data_quality': {}
        }
    
    def validate_file_exists(self) -> bool:
        """Check if Excel file exists."""
        if not os.path.exists(self.file_path):
            logger.error(f"✗ Excel file not found: {self.file_path}")
            return False
        
        logger.info(f"✓ Excel file found: {self.file_path}")
        self.structure_info['file_exists'] = True
        return True
    
    def load_workbook(self) -> bool:
        """Load Excel workbook."""
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            logger.info(f"✓ Loaded workbook with sheets: {', '.join(self.sheet_names)}")
            self.structure_info['sheets'] = {
                'count': len(self.sheet_names),
                'names': self.sheet_names
            }
            return True
        except Exception as e:
            logger.error(f"✗ Failed to load workbook: {e}")
            return False
    
    def validate_transactions_sheet(self) -> bool:
        """Validate that 'transactions ' sheet exists (note: trailing space)."""
        expected_sheet = "transactions "
        
        if expected_sheet not in self.sheet_names:
            logger.error(f"✗ Expected sheet '{expected_sheet}' not found")
            logger.info(f"  Available sheets: {', '.join(self.sheet_names)}")
            return False
        
        logger.info(f"✓ Found transactions sheet: '{expected_sheet}'")
        return True
    
    def read_transactions_sheet(self) -> bool:
        """Read transactions sheet with proper header handling."""
        try:
            # Read Excel with header in row 0 (Arabic headers)
            self.data = pd.read_excel(
                self.file_path,
                sheet_name="transactions ",
                header=0,
                dtype=str  # Read all as strings initially
            )
            
            logger.info(f"✓ Read transactions sheet: {len(self.data)} rows")
            return True
        except Exception as e:
            logger.error(f"✗ Failed to read transactions sheet: {e}")
            return False
    
    def validate_columns(self) -> Tuple[bool, List[str], List[str]]:
        """Validate that all expected columns are present."""
        if self.data is None:
            logger.error("✗ No data loaded")
            return False, [], []
        
        actual_columns = set(self.data.columns)
        expected_columns = set(EXPECTED_COLUMNS.keys())
        
        missing_columns = expected_columns - actual_columns
        extra_columns = actual_columns - expected_columns
        
        if missing_columns:
            logger.error(f"✗ Missing columns: {', '.join(missing_columns)}")
        else:
            logger.info(f"✓ All expected columns present")
        
        if extra_columns:
            logger.warning(f"⚠ Extra columns found: {', '.join(extra_columns)}")
        
        self.structure_info['validation_results']['columns'] = {
            'expected': len(expected_columns),
            'actual': len(actual_columns),
            'missing': list(missing_columns),
            'extra': list(extra_columns)
        }
        
        return len(missing_columns) == 0, list(missing_columns), list(extra_columns)
    
    def analyze_data_quality(self) -> Dict[str, Any]:
        """Analyze data quality issues."""
        if self.data is None:
            return {}
        
        quality_report = {
            'total_rows': len(self.data),
            'columns_analyzed': {}
        }
        
        for col in self.data.columns:
            col_data = self.data[col]
            null_count = col_data.isna().sum() + (col_data == '').sum()
            
            quality_report['columns_analyzed'][col] = {
                'total': len(col_data),
                'non_null': len(col_data) - null_count,
                'null_count': null_count,
                'null_percentage': round((null_count / len(col_data)) * 100, 2),
                'unique_values': col_data.nunique()
            }
            
            if null_count > 0:
                logger.warning(f"  {col}: {null_count} null values ({quality_report['columns_analyzed'][col]['null_percentage']}%)")
        
        self.structure_info['data_quality'] = quality_report
        return quality_report
    
    def analyze_data_types(self) -> Dict[str, str]:
        """Analyze data types in the Excel file."""
        if self.data is None:
            return {}
        
        type_analysis = {}
        for col in self.data.columns:
            type_analysis[col] = str(self.data[col].dtype)
        
        logger.info("Data types detected:")
        for col, dtype in type_analysis.items():
            logger.info(f"  {col}: {dtype}")
        
        self.structure_info['data_types'] = type_analysis
        return type_analysis
    
    def validate_structure(self) -> bool:
        """Run complete structure validation."""
        logger.info("=" * 60)
        logger.info("Validating Excel Structure")
        logger.info("=" * 60)
        
        # Step 1: File exists
        if not self.validate_file_exists():
            return False
        
        # Step 2: Load workbook
        if not self.load_workbook():
            return False
        
        # Step 3: Validate transactions sheet
        if not self.validate_transactions_sheet():
            return False
        
        # Step 4: Read data
        if not self.read_transactions_sheet():
            return False
        
        # Step 5: Validate columns
        columns_valid, missing, extra = self.validate_columns()
        if not columns_valid:
            logger.error("✗ Column validation failed")
            return False
        
        # Step 6: Analyze data quality
        logger.info("Analyzing data quality...")
        self.analyze_data_quality()
        
        # Step 7: Analyze data types
        logger.info("Analyzing data types...")
        self.analyze_data_types()
        
        logger.info("✓ Structure validation complete")
        return True
    
    def export_json(self, output_path: str) -> bool:
        """Export structure information to JSON."""
        try:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.structure_info, f, indent=2, ensure_ascii=False)
            logger.info(f"✓ Exported structure to: {output_path}")
            return True
        except Exception as e:
            logger.error(f"✗ Failed to export JSON: {e}")
            return False
    
    def generate_markdown_report(self, output_path: str) -> bool:
        """Generate human-readable Markdown report."""
        try:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("# Excel Structure Validation Report\n\n")
                f.write(f"**Generated**: {self.structure_info['timestamp']}\n\n")
                f.write(f"**File**: {self.structure_info['file_path']}\n\n")
                
                # Sheets section
                f.write("## Sheets\n\n")
                sheets_info = self.structure_info.get('sheets', {})
                f.write(f"- **Count**: {sheets_info.get('count', 0)}\n")
                f.write(f"- **Names**: {', '.join(sheets_info.get('names', []))}\n\n")
                
                # Columns section
                f.write("## Columns\n\n")
                col_validation = self.structure_info.get('validation_results', {}).get('columns', {})
                f.write(f"- **Expected**: {col_validation.get('expected', 0)}\n")
                f.write(f"- **Actual**: {col_validation.get('actual', 0)}\n")
                
                if col_validation.get('missing'):
                    f.write(f"- **Missing**: {', '.join(col_validation['missing'])}\n")
                
                if col_validation.get('extra'):
                    f.write(f"- **Extra**: {', '.join(col_validation['extra'])}\n")
                
                f.write("\n")
                
                # Data Quality section
                f.write("## Data Quality\n\n")
                quality = self.structure_info.get('data_quality', {})
                f.write(f"- **Total Rows**: {quality.get('total_rows', 0)}\n\n")
                
                f.write("### Column Analysis\n\n")
                f.write("| Column | Total | Non-Null | Null | Null % | Unique |\n")
                f.write("|--------|-------|----------|------|--------|--------|\n")
                
                for col, stats in quality.get('columns_analyzed', {}).items():
                    f.write(f"| {col} | {stats['total']} | {stats['non_null']} | {stats['null_count']} | {stats['null_percentage']}% | {stats['unique_values']} |\n")
                
                f.write("\n")
                
                # Data Types section
                f.write("## Data Types\n\n")
                f.write("| Column | Type |\n")
                f.write("|--------|------|\n")
                
                for col, dtype in self.structure_info.get('data_types', {}).items():
                    f.write(f"| {col} | {dtype} |\n")
                
                f.write("\n")
            
            logger.info(f"✓ Generated Markdown report: {output_path}")
            return True
        except Exception as e:
            logger.error(f"✗ Failed to generate Markdown: {e}")
            return False


def main():
    """Main execution function."""
    logger.info("=" * 60)
    logger.info("Phase 0 Task 0.2: Excel Structure Validation")
    logger.info("=" * 60)
    
    # Validate environment
    if not EXCEL_FILE_PATH:
        logger.error("✗ EXCEL_FILE_PATH not set in .env")
        return False
    
    # Create validator
    validator = ExcelStructureValidator(EXCEL_FILE_PATH)
    
    # Run validation
    if not validator.validate_structure():
        logger.error("\n" + "=" * 60)
        logger.error("✗ Phase 0 Task 0.2 FAILED")
        logger.error("=" * 60)
        return False
    
    # Export results
    json_path = "reports/excel_structure.json"
    md_path = "reports/excel_structure.md"
    
    success = True
    success = validator.export_json(json_path) and success
    success = validator.generate_markdown_report(md_path) and success
    
    if success:
        logger.info("\n" + "=" * 60)
        logger.info("✓ Phase 0 Task 0.2 COMPLETED")
        logger.info("=" * 60)
        logger.info(f"Structure exported to:")
        logger.info(f"  - {json_path}")
        logger.info(f"  - {md_path}")
        return True
    else:
        logger.error("\n" + "=" * 60)
        logger.error("✗ Phase 0 Task 0.2 FAILED")
        logger.error("=" * 60)
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
