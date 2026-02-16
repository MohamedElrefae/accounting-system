"""
Excel Reader for Excel Data Migration

This module provides enhanced Excel reading capabilities:
- Load column mapping from config/column_mapping_APPROVED.csv
- Read "transactions " sheet with proper header handling (skip row 0)
- Apply English column names automatically
- Return DataFrame with standardized column names
"""

import os
import json
import logging
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class ColumnMapping:
    """Column mapping from Excel to Supabase"""
    excel_column: str
    english_name: str
    supabase_table: str
    supabase_column: str
    data_type: str
    required: bool
    notes: Optional[str] = None


@dataclass
class ExcelStructure:
    """Structure information for Excel file"""
    file_path: str
    sheet_names: List[str]
    column_mappings: Dict[str, ColumnMapping]
    total_rows: int
    data_types: Dict[str, str]
    validation_errors: List[str] = field(default_factory=list)
    validation_warnings: List[str] = field(default_factory=list)


@dataclass
class ReadResult:
    """Result of Excel reading operation"""
    success: bool
    data: Optional[pd.DataFrame] = None
    structure: Optional[ExcelStructure] = None
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


class ExcelReader:
    """
    Enhanced Excel reader with column mapping support.
    
    This class reads Excel files with Arabic headers and applies
    English column names based on approved mapping configuration.
    """
    
    def __init__(self, excel_file_path: Optional[str] = None, 
                 mapping_file_path: Optional[str] = None):
        """
        Initialize Excel reader.
        
        Args:
            excel_file_path: Path to Excel file (optional, can be loaded from environment)
            mapping_file_path: Path to column mapping CSV (default: config/column_mapping_APPROVED.csv)
        """
        self.excel_file_path = excel_file_path or os.getenv("EXCEL_FILE_PATH")
        if not self.excel_file_path:
            raise ValueError("Excel file path not provided and EXCEL_FILE_PATH environment variable not set")
        
        if mapping_file_path is None:
            mapping_file_path = "config/column_mapping_APPROVED.csv"
        self.mapping_file_path = Path(mapping_file_path)
        
        self.workbook = None
        self.sheet_names = []
        self.column_mappings: Dict[str, ColumnMapping] = {}
        self.structure: Optional[ExcelStructure] = None
        
        # Load column mappings
        self.load_column_mappings()
    
    def load_column_mappings(self) -> bool:
        """
        Load column mappings from CSV file.
        
        Returns:
            True if mappings loaded successfully, False otherwise
        """
        try:
            if not self.mapping_file_path.exists():
                logger.error(f"Column mapping file not found: {self.mapping_file_path}")
                return False
            
            logger.info(f"Loading column mappings from {self.mapping_file_path}")
            
            # Read CSV file
            df_mappings = pd.read_csv(self.mapping_file_path)
            
            # Parse mappings
            for _, row in df_mappings.iterrows():
                # Get values with proper handling of NaN
                excel_col = str(row.get("Excel_Column", "")).strip() if pd.notna(row.get("Excel_Column")) else ""
                english_name = str(row.get("English_Name", "")).strip() if pd.notna(row.get("English_Name")) else ""
                supabase_table = str(row.get("Supabase_Table", "")).strip() if pd.notna(row.get("Supabase_Table")) else ""
                supabase_column = str(row.get("Supabase_Column", "")).strip() if pd.notna(row.get("Supabase_Column")) else ""
                data_type = str(row.get("Data_Type", "string")).strip() if pd.notna(row.get("Data_Type")) else "string"
                required_str = str(row.get("Required", "No")).strip().lower() if pd.notna(row.get("Required")) else "no"
                notes = str(row.get("Notes", "")).strip() if pd.notna(row.get("Notes")) else None
                
                mapping = ColumnMapping(
                    excel_column=excel_col,
                    english_name=english_name,
                    supabase_table=supabase_table,
                    supabase_column=supabase_column,
                    data_type=data_type,
                    required=required_str == "yes",
                    notes=notes if notes else None
                )
                
                # Store mapping by Excel column name
                if excel_col:
                    self.column_mappings[excel_col] = mapping
            
            logger.info(f"Loaded {len(self.column_mappings)} column mappings")
            return True
            
        except Exception as e:
            logger.error(f"Failed to load column mappings: {str(e)}")
            return False
    
    def validate_file_exists(self) -> bool:
        """
        Validate that Excel file exists.
        
        Returns:
            True if file exists, False otherwise
        """
        if not os.path.exists(self.excel_file_path):
            logger.error(f"Excel file not found: {self.excel_file_path}")
            return False
        
        logger.info(f"Excel file found: {self.excel_file_path}")
        return True
    
    def load_workbook(self) -> bool:
        """
        Load Excel workbook.
        
        Returns:
            True if workbook loaded successfully, False otherwise
        """
        try:
            self.workbook = load_workbook(self.excel_file_path, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            logger.info(f"Loaded workbook with sheets: {', '.join(self.sheet_names)}")
            return True
        except Exception as e:
            logger.error(f"Failed to load workbook: {str(e)}")
            return False
    
    def validate_transactions_sheet(self) -> bool:
        """
        Validate that 'transactions ' sheet exists (note: trailing space).
        
        Returns:
            True if sheet exists, False otherwise
        """
        expected_sheet = "transactions "
        
        if expected_sheet not in self.sheet_names:
            logger.error(f"Expected sheet '{expected_sheet}' not found")
            logger.info(f"Available sheets: {', '.join(self.sheet_names)}")
            return False
        
        logger.info(f"Found transactions sheet: '{expected_sheet}'")
        return True
    
    def read_transactions_sheet(self) -> ReadResult:
        """
        Read transactions sheet with proper header handling.
        
        Returns:
            ReadResult with data and structure information
        """
        result = ReadResult(success=False)
        
        try:
            # Step 1: Validate file exists
            if not self.validate_file_exists():
                result.errors.append(f"Excel file not found: {self.excel_file_path}")
                return result
            
            # Step 2: Load workbook
            if not self.load_workbook():
                result.errors.append("Failed to load Excel workbook")
                return result
            
            # Step 3: Validate transactions sheet
            if not self.validate_transactions_sheet():
                result.errors.append("Transactions sheet not found")
                return result
            
            # Step 4: Read Excel with proper header handling
            # Note: Row 0 contains Arabic headers, data starts from row 1
            logger.info("Reading transactions sheet with Arabic headers...")
            
            # Read the sheet with header in row 0 (Arabic headers)
            df_raw = pd.read_excel(
                self.excel_file_path,
                sheet_name="transactions ",
                header=0,  # Arabic headers are in row 0
                dtype=str  # Read all as strings initially to preserve data
            )
            
            logger.info(f"Raw data read: {len(df_raw)} rows, {len(df_raw.columns)} columns")
            
            # Step 5: Apply English column names
            df_english = self._apply_english_column_names(df_raw)
            
            # Step 6: Create structure information
            self.structure = self._create_structure_info(df_english)
            
            # Step 7: Validate required columns
            validation_result = self._validate_required_columns(df_english)
            if not validation_result["success"]:
                result.errors.extend(validation_result["errors"])
                result.warnings.extend(validation_result["warnings"])
            
            # Update result
            result.success = True
            result.data = df_english
            result.structure = self.structure
            result.warnings.extend(validation_result["warnings"])
            
            logger.info(f"Successfully read {len(df_english)} rows with {len(df_english.columns)} English columns")
            
        except Exception as e:
            logger.error(f"Failed to read transactions sheet: {str(e)}")
            result.errors.append(f"Error reading Excel: {str(e)}")
        
        return result
    
    def _apply_english_column_names(self, df_raw: pd.DataFrame) -> pd.DataFrame:
        """
        Apply English column names to DataFrame.
        
        Args:
            df_raw: DataFrame with Arabic column names
            
        Returns:
            DataFrame with English column names
        """
        # Create a copy to avoid modifying the original
        df_english = df_raw.copy()
        
        # Build mapping from Arabic to English column names
        column_mapping = {}
        missing_mappings = []
        
        for arabic_col in df_raw.columns:
            # Clean column name (remove extra whitespace)
            arabic_col_clean = str(arabic_col).strip()
            
            # Find mapping - try exact match first, then try with spaces stripped from mapping keys
            mapping = self.column_mappings.get(arabic_col_clean)
            
            # If not found, try matching against stripped versions of all mapping keys
            if not mapping:
                for map_key, map_obj in self.column_mappings.items():
                    if map_key.strip() == arabic_col_clean:
                        mapping = map_obj
                        break
            
            if mapping:
                column_mapping[arabic_col] = mapping.english_name
                logger.debug(f"Mapped '{arabic_col}' -> '{mapping.english_name}'")
            else:
                # No mapping found, keep original name
                column_mapping[arabic_col] = arabic_col_clean
                missing_mappings.append(arabic_col_clean)
                logger.warning(f"No mapping found for column: '{arabic_col_clean}'")
        
        # Apply column mapping
        df_english = df_english.rename(columns=column_mapping)
        
        # Log missing mappings
        if missing_mappings:
            logger.warning(f"Missing mappings for {len(missing_mappings)} columns: {', '.join(missing_mappings)}")
        
        return df_english
    
    def _create_structure_info(self, df_english: pd.DataFrame) -> ExcelStructure:
        """
        Create structure information for Excel file.
        
        Args:
            df_english: DataFrame with English column names
            
        Returns:
            ExcelStructure object
        """
        # Analyze data types
        data_types = {}
        for col in df_english.columns:
            # Try to infer data type
            sample_value = df_english[col].iloc[0] if len(df_english) > 0 else None
            if pd.isna(sample_value):
                data_types[col] = "unknown"
            elif isinstance(sample_value, (int, float)):
                data_types[col] = "numeric"
            elif isinstance(sample_value, str):
                # Check if it looks like a date
                try:
                    pd.to_datetime(sample_value)
                    data_types[col] = "date"
                except (ValueError, TypeError):
                    data_types[col] = "string"
            else:
                data_types[col] = str(type(sample_value).__name__)
        
        # Create structure
        structure = ExcelStructure(
            file_path=self.excel_file_path,
            sheet_names=self.sheet_names,
            column_mappings=self.column_mappings,
            total_rows=len(df_english),
            data_types=data_types
        )
        
        return structure
    
    def _validate_required_columns(self, df_english: pd.DataFrame) -> Dict[str, Any]:
        """
        Validate that all required columns are present.
        
        Args:
            df_english: DataFrame with English column names
            
        Returns:
            Dictionary with validation results
        """
        result = {
            "success": True,
            "errors": [],
            "warnings": []
        }
        
        # Get required columns from mappings
        required_columns = []
        for mapping in self.column_mappings.values():
            if mapping.required and mapping.english_name:
                required_columns.append(mapping.english_name)
        
        # Check for missing required columns
        missing_columns = []
        for required_col in required_columns:
            if required_col not in df_english.columns:
                missing_columns.append(required_col)
        
        if missing_columns:
            result["success"] = False
            error_msg = f"Missing required columns: {', '.join(missing_columns)}"
            result["errors"].append(error_msg)
            logger.error(error_msg)
        
        # Check for null values in required columns
        for required_col in required_columns:
            if required_col in df_english.columns:
                null_count = df_english[required_col].isna().sum()
                if null_count > 0:
                    warning_msg = f"Column '{required_col}' has {null_count} null values"
                    result["warnings"].append(warning_msg)
                    logger.warning(warning_msg)
        
        return result
    
    def get_column_mapping(self, excel_column: str) -> Optional[ColumnMapping]:
        """
        Get column mapping for a specific Excel column.
        
        Args:
            excel_column: Excel column name (Arabic)
            
        Returns:
            ColumnMapping if found, None otherwise
        """
        return self.column_mappings.get(excel_column)
    
    def get_english_column_name(self, excel_column: str) -> Optional[str]:
        """
        Get English column name for a specific Excel column.
        
        Args:
            excel_column: Excel column name (Arabic)
            
        Returns:
            English column name if mapping exists, None otherwise
        """
        mapping = self.get_column_mapping(excel_column)
        return mapping.english_name if mapping else None
    
    def get_supabase_mapping(self, english_column: str) -> Optional[Tuple[str, str]]:
        """
        Get Supabase table and column for an English column name.
        
        Args:
            english_column: English column name
            
        Returns:
            Tuple of (supabase_table, supabase_column) if found, None otherwise
        """
        for mapping in self.column_mappings.values():
            if mapping.english_name == english_column:
                return (mapping.supabase_table, mapping.supabase_column)
        return None
    
    def export_structure_json(self, output_path: str) -> bool:
        """
        Export structure information to JSON file.
        
        Args:
            output_path: Path to output JSON file
            
        Returns:
            True if export successful, False otherwise
        """
        if not self.structure:
            logger.error("No structure information available")
            return False
        
        try:
            # Prepare structure data
            structure_data = {
                "timestamp": datetime.now().isoformat(),
                "file_path": self.structure.file_path,
                "sheet_names": self.structure.sheet_names,
                "total_rows": self.structure.total_rows,
                "data_types": self.structure.data_types,
                "column_mappings": [],
                "validation_errors": self.structure.validation_errors,
                "validation_warnings": self.structure.validation_warnings
            }
            
            # Add column mappings
            for mapping in self.column_mappings.values():
                mapping_data = {
                    "excel_column": mapping.excel_column,
                    "english_name": mapping.english_name,
                    "supabase_table": mapping.supabase_table,
                    "supabase_column": mapping.supabase_column,
                    "data_type": mapping.data_type,
                    "required": mapping.required,
                    "notes": mapping.notes
                }
                structure_data["column_mappings"].append(mapping_data)
            
            # Write to file
            output_path_obj = Path(output_path)
            output_path_obj.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path_obj, 'w', encoding='utf-8') as f:
                json.dump(structure_data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"Structure exported to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export structure: {str(e)}")
            return False
    
    def __str__(self) -> str:
        """String representation of Excel reader"""
        return f"ExcelReader(file='{self.excel_file_path}', mappings={len(self.column_mappings)})"


# Factory function for easy creation
def create_excel_reader(excel_file_path: Optional[str] = None, 
                       mapping_file_path: Optional[str] = None) -> ExcelReader:
    """
    Factory function to create Excel reader.
    
    Args:
        excel_file_path: Path to Excel file
        mapping_file_path: Path to column mapping CSV
        
    Returns:
        ExcelReader instance
    """
    return ExcelReader(excel_file_path, mapping_file_path)